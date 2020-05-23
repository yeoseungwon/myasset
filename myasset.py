import pandas as pd
import pandas_datareader.data as web
import matplotlib.pyplot as plt
from copy import copy
import numpy as np
import math
import random
import os
from datetime import datetime
from statsmodels import regression
import statsmodels.api as sm
from scipy import stats
from openpyxl import Workbook


def cal_mdd(x):
    arr_v = np.array(x)
    peak_lower = np.argmax(np.maximum.accumulate(arr_v) - arr_v)
    peak_upper = np.argmax(arr_v[:peak_lower])
    return (arr_v[peak_lower] - arr_v[peak_upper]) / arr_v[peak_upper] * 100


def cal_cagr(_data, ref, _day):
    _day = _day/240
    _cagr = math.pow(_data/ref, 1/_day)-1
    return _cagr * 100


class Stock:
    def __init__(self, stock_name, u, v):
        self.u = u    # 연 평균 증가율
        self.v = v    # 연 평균 변화율
        self.data = web.DataReader(stock_name, 'yahoo', '2007-02-01').Close[:]
        self.money = 0
        self.stock_num = 0
        self.stock_price = 0.0
        self.data_length = len(self.data)
        self.interest = 0.03
        self.d = pd.Series(self.data).to_numpy()
        self.momentum = 0.0
        self.checker = 0
        self.vol = 0.0
        self.mod_momentum = 0.0
        self.hold_day = 0
        self.today = 0
        self.stock_name = stock_name

    def set_montecarlo(self):  # 기하 브라운 운동
        arr = np.zeros(self.data_length)
        arr[0] = self.d[0]
        u0 = self.u / 240
        vol0 = self.v / math.sqrt(240)
        for ii in range(1, self.data_length):
            arr[ii] = arr[ii - 1] * math.exp(u0 - 0.5 * vol0 * vol0 + vol0 * random.gauss(0.0, 1.0))
        self.data = arr     # 몬테카를로 시작하면 data에 임의의 주가가 들어가있음
        self.d = pd.Series(self.data).to_numpy()

    def cal_momentum_score(self, _day):  # 그 날의 모멘텀 스코어 계산
        self.momentum = 0.0
        temp0 = 1.0
        temp1 = 1.0
        temp2 = 1.2
        temp3 = 1.3
        for ii in range(1, 13):
            if ii < 4:
                temp = copy(temp0)
            elif (ii >= 4) and (ii < 7):
                temp = copy(temp1)
            elif (ii >= 7) and (ii < 10):
                temp = copy(temp2)
            else:
                temp = copy(temp3)
            self.momentum += temp * (self.d[_day] // (self.d[_day - ii * 20] * (1 + self.interest * ii / 12)))
        self.momentum = self.momentum / (temp0 * 3 + temp1 * 3 + temp2 * 3 + temp3 * 3)
        return self.momentum

    def trading(self, ss_rate, _day):  # ss_rate 는 내가 보유해야 하는 개수
        self.money = 0
        if self.stock_name == 'GLD':
            self.checker = 1

        if self.stock_num > ss_rate:  # 매도, 매도는 평단가 현재가 보다 높을 때만 진행
            if self.stock_price * 1.008 <= self.d[_day] or self.checker == 1:
                self.money = (self.stock_num - ss_rate) * self.d[_day] * 0.997
                self.stock_num = ss_rate
                if self.stock_num == 0:
                    self.stock_price = 0
                else:
                    pass
            elif self.stock_price > self.d[_day] and self.checker == 0:  # 현재가보다 평단가가 낮아서 일부만 매도
                if round(self.stock_num * 1) > ss_rate:
                    ss_rate = self.stock_num * 1
                else:
                    pass
                self.money = (self.stock_num - ss_rate) * self.d[_day] * 0.997
                self.stock_num = ss_rate
                if self.stock_num == 0:
                    self.stock_price = 0
                else:
                    pass
            else:
                pass
            if self.stock_num == 0:
                self.hold_day = 0

        else:  # 매수
            self.money = (self.stock_num - ss_rate) * self.d[_day] * 1.003
            if ss_rate == 0:
                self.money = 0
            else:
                self.stock_price = (self.stock_price * self.stock_num + (ss_rate - self.stock_num) * self.d[_day]) \
                                   / ss_rate
                self.stock_num = ss_rate
            self.hold_day = _day
        return self.money

    def check_losscut(self, _day):
        if _day == self.hold_day:
            return 0
        cl_a = (- self.stock_price + self.d[_day]) / self.stock_price
        cl_b = (self.u / 240 * (_day - self.hold_day) - (self.v / math.sqrt(240 / (_day - self.hold_day)) * 3000.1))
        if cl_a < cl_b:
            self.checker = 1
            return 1
        else:
            return 0

    def cal_mod_vol(self, _day):
        self.vol = 0.0
        temp_v = []
        for ii in range(0, 12):
            temp_v.append((self.d[_day - ii * 20] - self.d[_day - ii * 20 - 20]) / self.d[_day - ii * 20 - 20] * 100)
        self.vol = np.std(temp_v)
        return self.vol

    def cal_mod_ms(self, _day):
        self.mod_momentum = 0.0
        temp0 = 1.0
        temp1 = 1.0
        temp2 = 1.2
        temp3 = 1.5
        for ii in range(1, 13):
            if ii < 4:
                temp = copy(temp0)
            elif (ii >= 4) and (ii < 7):
                temp = copy(temp1)
            elif (ii >= 7) and (ii < 10):
                temp = copy(temp2)
            else:
                temp = copy(temp3)
            self.mod_momentum += temp * (self.d[_day] / (self.d[_day - ii * 20] * (1 + self.interest * ii / 12)))
        self.mod_momentum = self.mod_momentum / (temp0 * 3 + temp1 * 3 + temp2 * 3 + temp3 * 3)
        return self.mod_momentum


# 시스템 손절매
def cal_momen_losscut(asset, _day):
    momentum = 0.0
    temp0 = 1.0
    temp1 = 2.0
    temp = 0.0
    momen_rapid_lc = 1
    _vol = []
    for i_lc in range(1, 28):
        if i_lc < 15:
            temp = copy(temp0)
        else:
            temp = copy(temp1)
        momentum += (asset[_day] * 1.04 // asset[_day - i_lc * 5]) * temp
    momentum = momentum / (temp0 * 14 + temp1 * 13)
    # 급락 방어
    if _day > 243:
        for i_lc in range(242, _day - 1):
            _vol.append((asset[i_lc] - asset[i_lc - 1]) / asset[i_lc - 1])
        vol_val = np.std(_vol)
        avg_val = np.average(_vol)
        if (asset[_day] - asset[_day - 1]) / asset[i_lc - 1] < avg_val - 3 * vol_val:
            momen_rapid_lc = 0.5
        momentum = momen_rapid_lc * momentum

    return momentum


def linreg(x, y):
    x = sm.add_constant(x)
    model = regression.linear_model.OLS(y, x).fit()

    x = x[:, 1]
    return model.params[0], model.params[1]

today_price_checker = 0
now_price = [1000.0, 1000.0, 1000.0, 1000.0, 1000.0]
year = datetime.today().year
month = datetime.today().month
day = datetime.today().day
today = pd.Timestamp(year, month, day, 00, 00, 00)

# my_stock_name = ['VYM', 'TLT', 'QQQ', 'VDC', 'IAU']  # 새로운  # 기존
my_stock_name = ['VYM', 'TLT', 'QQQ', 'VDC', 'GLD']  # 새로운
stock_ref = 'SPY'
my_u = [0.0624, 0.0745, 0.0509, 0.0912, 0.0825]
my_v = [0.1485, 0.1314, 0.2361, 0.1113, 0.1759]
my_stock = []
momen = []
smomen = []
s_rate = []
temp_s = []
sim_n = 1
save_mc_result = np.zeros((sim_n, 2))
add_money = 0
cash_rate = 0.5
start_money = 8000
real_money = 8000
avg_MDD = 0.0
avg_CAGR = 0.0
# write_wb = Workbook()
# write_ws = write_wb.active
lc_momen = 0.0

print("시뮬레이션용 코드")
print("Data 입력 시작")
stocks = len(my_stock_name)

debug_i = 0

for j in range(stocks):
    my_stock.append(Stock(my_stock_name[j], my_u[j], my_v[j]))
    print(j+1, "번 데이터 입력 완료")
    momen.append(0.0)
    smomen.append(0.0)
    s_rate.append(0.0)
    temp_s.append(0.0)
    if today_price_checker == 1:
        if my_stock[j].data.index[-1] == today:
            my_stock[j].data[-1] = now_price[j]
        else:
            my_stock[j].data.loc[today] = now_price[j]
    else:
        pass
    my_stock[j].data_length = len(my_stock[j].data)
    my_stock[j].d = pd.Series(my_stock[j].data).to_numpy()
# 비교 기준
ref_my_stock = Stock(stock_ref, 0, 0)
ref_my_money = copy(start_money)
ref_My_Asset = np.zeros(my_stock[0].data_length)
if today_price_checker == 1:
    if ref_my_stock.data.index[-1] == today:
        ref_my_stock.data[-1] = ref_my_stock.data[-1]
    else:
        ref_my_stock.data.loc[today] = ref_my_stock.data[-1]
else:
    pass
ref_my_stock.data_length = len(ref_my_stock.data)
ref_my_stock.d = pd.Series(ref_my_stock.data).to_numpy()

bank_My_Asset = np.zeros(my_stock[0].data_length)
bank_My_Asset[0] = copy(start_money)
for i in range(1, my_stock[0].data_length):
    if i >= 240:
        bank_My_Asset[i] = bank_My_Asset[i-1] * (1 + 0.03 / 240)
        if i % 20 == (my_stock[0].data_length - 1) % 20:
            bank_My_Asset[i] += add_money
        else:
            pass
    else:
        bank_My_Asset[i] = bank_My_Asset[i-1]
        if i % 20 == (my_stock[0].data_length - 1) % 20:
            bank_My_Asset[i] += add_money
        else:
            pass

cal_for_cagr = start_money + add_money * (my_stock[0].data_length - 240) // 20

for k in range(0, sim_n):
    My_Asset = np.zeros(my_stock[0].data_length)
    my_money = copy(start_money)
    if k > 0:
        for j in range(stocks):
            my_stock[j].set_montecarlo()
            my_stock[j].stock_num = 0
            my_stock[j].stock_price = 0
            my_stock[j].hold_day = 0
    else:
        pass

    for i in range(my_stock[0].data_length):

        if i == my_stock[0].data_length-1:
            for j in range(stocks):
                temp_s[j] = my_stock[j].stock_num
        else:
            pass
        for j in range(stocks):  # 모멘텀 계산
            if i < 240:
                momen[j] = 0
            else:
                # momen[j] = my_stock[j].cal_momentum_score(i)
                momen[j] = my_stock[j].cal_mod_ms(i)
                ref_my_stock.cal_momentum_score(i)

        if i >= 240:
            if i % 20 == (my_stock[0].data_length - 1) % 20:
                my_money += add_money
                ref_my_money += add_money
            else:
                pass

            if i % 20 == (my_stock[0].data_length - 1) % 20:  # 한달에 한번 모멘텀 스코어 이용해서 매매, 상대 모멘텀으로 주식 종목수/1.3 선정
                ref_my_stock.stock_num += ref_my_money // ref_my_stock.d[i]
                ref_my_money -= ref_my_money // ref_my_stock.d[i] * ref_my_stock.d[i]

                c_momen = copy(momen)
                c_momen.sort()
                c_momen.reverse()
                temp_money = copy(my_money)
                # 상대 모멘텀
                for j in range(stocks):
                    smomen[j] = 0.00
                for j in range(3):  # 상대 모멘텀으로 상위 x% 선정, smomen에 상위 x% 주식 번호 저장
                    for jj in range(stocks):
                        if momen[jj] == c_momen[j]:
                            smomen[jj] = 1.0
                            break
                        else:
                            pass
                # 상대 모멘텀 종료
                # 모멘텀 비율 이용하여 현금 비중 구하기
                cash_rate = 0.25
                temp_cash_rate = 0
                for j in range(stocks):
                    if smomen[j] == 1:
                        temp_cash_rate += 1
                        if momen[j] < 1:
                            cash_rate += 1
                        else:
                            pass
                    else:
                        momen[j] = 0
                for j in range(stocks):
                    momen[j] = momen[j] / my_stock[j].cal_mod_vol(i)
                cash_rate = cash_rate / temp_cash_rate * 0.2

                sum_momen = 0.0
                t_temp_money = copy(my_money)
                for j in range(stocks):
                    temp_money += my_stock[j].d[i] * my_stock[j].stock_num
                    sum_momen += momen[j]
                sum_momen += cash_rate
                temp_checker = []
                temp_each_asset = []
                for j in range(stocks):
                    temp_each_asset.append(my_stock[j].d[i] * my_stock[j].stock_num)
                    temp_checker.append(1)
                while 1:
                    checker = 0
                    for j in range(stocks):
                        if temp_each_asset[j] > (sum(temp_each_asset) + my_money) * momen[j] / sum_momen and \
                                my_stock[j].d[i] < my_stock[j].stock_price * 1.008:
                            temp_checker[j] = 0
                            checker = 1
                        else:
                            temp_checker[j] = 1
                    if checker == 0:
                        break
                    else:
                        for j in range(stocks):
                            temp_each_asset[j] = temp_checker[j] * temp_each_asset[j]
                temp_money = sum(temp_each_asset) + my_money

                for j in range(stocks):
                    my_stock[j].checker = 0
                    momen[j] = math.floor(momen[j] * 100) / 100
                    s_rate[j] = math.floor((momen[j] / sum_momen * temp_money) / my_stock[j].d[i])
                    # 값에 따라 매수 또는 매도 진행
                    if j == stocks - 1:
                        s_rate[j] = s_rate[j] * 0.2
                    my_money += my_stock[j].trading(s_rate[j], i)

                if my_money < 0:
                    print(i)
                    print("WTF")
                else:
                    pass
            # 종목별 익절매
            for j in range(stocks - 1):
                if my_stock[j].stock_price * 1.15 < my_stock[j].d[i] and my_stock[j].stock_price != 0:
                    my_stock[j].checker = 1
                    my_money += my_stock[j].trading(math.floor(my_stock[j].stock_num * 0.0), i)
                    my_stock[j].checker = 0
                else:
                    pass
            if my_stock[stocks - 1].stock_price * 1.1 < my_stock[stocks - 1].d[i] and my_stock[stocks - 1].stock_price != 0:
                my_stock[stocks - 1].checker = 1
                my_money += my_stock[stocks - 1].trading(math.floor(my_stock[stocks - 1].stock_num * 0.0), i)
                my_stock[stocks - 1].checker = 0
            else:
                pass

            if my_money < 0:
                print("W")
            else:
                pass

            # 시스템 손절매
            for j in range(stocks):
                My_Asset[i] += my_stock[j].d[i] * my_stock[j].stock_num
            My_Asset[i] += my_money
            lc_momen = cal_momen_losscut(My_Asset, i)
            if lc_momen < (1 - my_money / My_Asset[i]):
                lc_momen = lc_momen / (1 - my_money / My_Asset[i])
                for j in range(stocks):
                    my_stock[j].checker = 0
                    s_rate[j] = round(lc_momen * my_stock[j].stock_num)
                    my_money += my_stock[j].trading(s_rate[j], i)
            else:
                pass

        else:
            if i % 20 == (my_stock[0].data_length - 1) % 20:
                my_money += add_money
                ref_my_money += add_money
            else:
                pass
        My_Asset[i] = 0
        ref_My_Asset[i] = 0
        for j in range(stocks):
            My_Asset[i] += my_stock[j].d[i] * my_stock[j].stock_num
        ref_My_Asset[i] += ref_my_stock.d[i] * ref_my_stock.stock_num + ref_my_money
        My_Asset[i] += my_money
        # TODO 엑셀로 데이터 출력하기 주식 수량, 구매가, 현재가, 현재 돈
        '''
        if k == 0:
            for j in range(stocks):
                write_ws.cell(i + 1, j * 3 + 1, my_stock[j].stock_num)
                write_ws.cell(i + 1, j * 3 + 2, my_stock[j].stock_price)
                write_ws.cell(i + 1, j * 3 + 3, my_stock[j].d[i])
            write_ws.cell(i + 1, 16, my_money)
            write_ws.cell(i + 1, 17, lc_momen)
'''
    if k == 0:
        mdd = cal_mdd(ref_My_Asset)
        print("REF MDD : ", '%.2f' % mdd, end='\t')
        cagr = cal_cagr(ref_My_Asset[-1], cal_for_cagr, my_stock[0].data_length)
        print("REF CAGR : ", '%.2f' % cagr)

        mdd = cal_mdd(My_Asset)
        print("TEST MDD : ", '%.2f' % mdd, end='\t')
        cagr = cal_cagr(My_Asset[-1], cal_for_cagr, my_stock[0].data_length)
        print("TEST CAGR : ", '%.2f' % cagr)

        My_Asset_plot = pd.Series(My_Asset, index=my_stock[0].data.index)
        ref_My_Asset_plot = pd.Series(ref_My_Asset, index=my_stock[0].data.index)
        bank_My_Asset_plot = pd.Series(bank_My_Asset, index=my_stock[0].data.index)

        plt.subplot()
        plt.plot(ref_My_Asset_plot, label='ref')
        plt.plot(bank_My_Asset_plot, label='BANK')
        plt.plot(My_Asset_plot, label='TEST')

        print("주식 보유량 및 변화량")
        # rate = real_money / My_Asset[-1]
        for j in range(stocks):
            print(my_stock_name[j], end='\t')
            print("보유량 : ", '%3.f' % my_stock[j].stock_num, end='\t')
            print("변화량 : ", '%3.f' % (my_stock[j].stock_num - temp_s[j]), end='\t')
            print("Momentum : ", '%.2f' % momen[j])
        print("CASH : ", '%.2f' % my_money)
        print("총 자산 : ", '%.2f' % My_Asset[-1])
        print("시스템 손절매 모멘텀:", '%.2f' % lc_momen)

        # _spy = web.DataReader('SPY', 'yahoo', '2007-02-01')
        # benchmark_spy = _spy.Close.pct_change()[1:]
        benchmark_ref = ref_My_Asset_plot.pct_change()[1:]
        benchmark_asset = My_Asset_plot.pct_change()[1:]
        alpha, beta = linreg(benchmark_ref.values[260:], benchmark_asset.values[260:])
        # (beta, alpha) = stats.linregress(benchmark_spy.values[260:], benchmark_asset.values[260:])[0:2]  # 위와 동일함
        asset_pct_change = benchmark_asset.to_numpy()
        asset_stdev = np.std(asset_pct_change) * math.sqrt(240)
        sharpratio = (cagr / 100 - 0.03) / asset_stdev
        print("sharp:", '%.5f' % sharpratio)
        print("beta:", '%.5f' % beta)
    else:
        pass
    if k % 1 == 0:
        print(k)
    mdd = cal_mdd(My_Asset)
    save_mc_result[k, 0] = mdd
    cagr = cal_cagr(My_Asset[-1], cal_for_cagr, my_stock[0].data_length)
    save_mc_result[k, 1] = cagr
    avg_MDD = (avg_MDD * k + mdd) / (k + 1)
    avg_CAGR = (avg_CAGR * k + cagr) / (k + 1)

print("AVG MDD : ", '%.2f' % avg_MDD, end='\t')
print("AVG CAGR : ", '%.2f' % avg_CAGR)

# df = pd.DataFrame(save_mc_result)
data = {'mdd': save_mc_result[:, 0], 'CAGR': save_mc_result[:, 1]}
df = pd.DataFrame(data)
df.to_excel("output.xlsx")

plt.legend(loc='best')
plt.grid()
plt.show()
#write_wb.save('전체값.xlsx')
print("end")
